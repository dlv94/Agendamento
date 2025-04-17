import os
import winsound
from time import sleep
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from utils import iniciar_driver


def executar_agendamentos(quadra, horarios):
    planilha = os.path.join(os.path.dirname(__file__), 'planilha.xlsx')
    workbook = load_workbook(filename=planilha)
    sheet = workbook['dados']

    driver, wait = iniciar_driver()

    cpfs_cadastrados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        cpf, nome, email, telefone = row
        if cpf in cpfs_cadastrados:
            continue

        sucesso = realizar_agendamento(driver, wait, cpf, nome, email, telefone, quadra, horarios)
        if sucesso:
            cpfs_cadastrados.append(cpf)


def realizar_agendamento(driver, wait, cpf, nome, email, telefone, quadra, horarios):
    try:
        driver.get('https://agendamento.americana.sp.gov.br/esportes/agenda-index.php?a=identificacao')
        driver.execute_script("document.body.style.zoom='65%'")

        wait.until(EC.presence_of_element_located((By.ID, "txtCpf"))).send_keys(cpf)
        wait.until(EC.presence_of_element_located((By.ID, "txtNome"))).send_keys(nome)
        wait.until(EC.presence_of_element_located((By.ID, "txtEmail"))).send_keys(email)

        fone_field = wait.until(EC.presence_of_element_located((By.ID, "txtFone")))
        fone_field.click()
        sleep(0.2)
        for n in str(telefone):
            fone_field.send_keys(n)

        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/form/div[6]/button[2]'))).click()

        print(f"Realizando agendamento para CPF:{cpf} - Local:{quadra} - Horário:{horarios}")

        if quadra == "Areia Meneghel":
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="cboEquipamento"]/optgroup[3]/option[3]'))).click()
        elif quadra == "Quadra Meneghel":
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="cboEquipamento"]/optgroup[2]/option[4]'))).click()
        elif quadra == "Quadra Nova Americana":
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="cboEquipamento"]/optgroup[2]/option[3]'))).click()

        driver.execute_script("document.body.style.zoom='65%'")

        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/form/div[4]/button[3]'))).click()

        sleep(0.5)  # Espera carregamento do calendário

        dia = datetime.now().day
        xpath_base = '/html/body/div[1]/div[2]/div/form/div[1]/div[1]/div/div[1]/table/tbody/tr'

        driver.execute_script("document.body.style.zoom='65%'")

        for semana in range(1, 7):
            for dia_semana in range(1, 8):
                try:
                    celula = wait.until(EC.presence_of_element_located((By.XPATH, f'{xpath_base}[{semana}]/td[{dia_semana}]')))
                    print("celula", celula.text.strip(),celula.get_attribute("class"))
                    if celula.text.strip() == str(dia) and 'disabled' not in celula.get_attribute("class"):
                        #if "disabled" not in celula.get_attribute("class"):
                        celula.click()
                        break
                except:
                    
                    continue

        # Verifica horários disponíveis e agenda
        for i in range(1, 25):
            try:
                opcao = driver.find_element(By.XPATH, f'/html/body/div[1]/div[2]/div/form/div[1]/div[2]/select/option[{i}]')
                print("eis op", opcao.text)
                if opcao.text in horarios:
                    opcao.click()
                    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/form/div[3]/button[3]'))).click()

                    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/form/div[7]/div/input'))).click()
                    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/form/div[8]/div/input'))).click()
                    sleep(1.2)

                    winsound.PlaySound("SystemHand", winsound.SND_ASYNC)

                    return True
            except:
                continue

        return False

    except Exception as e:
        print(f"Erro ao tentar agendar para {cpf}: {e}")
        return False